import openpyxl
from openpyxl.chart import BarChart, Reference

wookbook = openpyxl.load_workbook('transactions.xlsx')

# lire le fichier actif
# sheet = wookbook.active

# on parcours le fichier
# for i in range(0, sheet.max_row):
#   for col in sheet.iter_cols(1, sheet.max_column):
#      print(col[i].value, end="\t\t")
# print('')

sheet = wookbook['Sheet1']

# get a cel
cell = sheet['a2']
print(cell.value)
print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
    # price with discount
    price_width_discount = cell.value * 0.9
    fourth_cell = sheet.cell(row, 4)
    fourth_cell.value = price_width_discount

values = Reference(
    sheet,
    min_row=2,
    max_row=sheet.max_row,
    min_col=4,
    max_col=4
)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wookbook.save('transactions2.xlsx')
